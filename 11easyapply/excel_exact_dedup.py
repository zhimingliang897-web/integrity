#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 整行完全一致去重脚本

规则：
- 只有“整行每个单元格都一致”才判定为重复
- 重复行只删除后续副本，保留第一条
- 只要有任意一个单元格不同，就保留

默认行为：
- 默认把每个工作表第 1 行当表头保留，不参与去重
- 默认处理所有工作表
"""

from __future__ import annotations

import argparse
import os
import sys
from typing import Iterable, Tuple

try:
    from openpyxl import load_workbook
except Exception:
    print("缺少依赖：openpyxl")
    print("请先安装：pip install openpyxl")
    sys.exit(1)


def normalize_row(row_values: Iterable[object]) -> Tuple[object, ...]:
    """将一行值转成可哈希的 tuple，用于精确比较。"""
    return tuple(row_values)


def dedup_worksheet(ws, has_header: bool = True) -> tuple[int, int]:
    """
    对单个工作表去重（原地改写）。
    返回：(原始行数, 删除行数)
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, 0

    start_idx = 1 if has_header else 0
    kept_rows = rows[:start_idx]
    seen = set()
    removed = 0

    for row in rows[start_idx:]:
        key = normalize_row(row)
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        kept_rows.append(row)

    # 清空并重写内容（仅处理值，不保留样式）
    ws.delete_rows(1, ws.max_row)
    for new_row_idx, row in enumerate(kept_rows, start=1):
        for new_col_idx, value in enumerate(row, start=1):
            ws.cell(row=new_row_idx, column=new_col_idx, value=value)

    return len(rows), removed


def build_default_output_path(input_path: str) -> str:
    base, ext = os.path.splitext(input_path)
    return f"{base}_去重{ext or '.xlsx'}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Excel 按整行完全一致去重（相同行仅保留第一条）。"
    )
    parser.add_argument("input", help="输入 Excel 文件路径（.xlsx）")
    parser.add_argument(
        "-o",
        "--output",
        help="输出文件路径（默认：在原文件名后加 _去重）",
    )
    parser.add_argument(
        "--sheet",
        help="只处理指定工作表（默认处理全部工作表）",
    )
    parser.add_argument(
        "--no-header",
        action="store_true",
        help="不保留首行表头（即首行也参与去重）",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = args.input
    output_path = args.output or build_default_output_path(input_path)
    has_header = not args.no_header

    if not os.path.exists(input_path):
        print(f"输入文件不存在：{input_path}")
        return 1

    if not input_path.lower().endswith(".xlsx"):
        print("当前脚本只支持 .xlsx 文件")
        return 1

    wb = load_workbook(input_path)
    sheets = [wb[args.sheet]] if args.sheet else wb.worksheets

    total_before = 0
    total_removed = 0

    for ws in sheets:
        before, removed = dedup_worksheet(ws, has_header=has_header)
        total_before += before
        total_removed += removed
        print(f"[{ws.title}] 原始 {before} 行，删除重复 {removed} 行")

    wb.save(output_path)
    print("-" * 40)
    print(f"处理完成：{input_path}")
    print(f"输出文件：{output_path}")
    print(f"总删除重复行：{total_removed}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

